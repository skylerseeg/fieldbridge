"""
Excel export for Buck's outreach list.
Generates a formatted .xlsx with one row per prospect, sorted by fit score.
Includes contact info, outreach angles, and score breakdown.

Usage:
    from fieldbridge.saas.prospect_intelligence.export import export_bucks_list
    path = export_bucks_list(prospects, contacts_by_prospect)
"""
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

log = logging.getLogger("fieldbridge.export")

# Score tier colors
_RED = "FFC7CE"       # < 40 — likely not a fit
_YELLOW = "FFEB9C"    # 40–69 — maybe
_GREEN = "C6EFCE"     # 70–84 — good fit
_TEAL = "92D050"      # 85+ — hot lead

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79") if HAS_OPENPYXL else None
_HEADER_FONT = Font(color="FFFFFF", bold=True) if HAS_OPENPYXL else None

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN) if HAS_OPENPYXL else None


def _score_fill(score: int):
    if score >= 85:
        color = _TEAL
    elif score >= 70:
        color = _GREEN
    elif score >= 40:
        color = _YELLOW
    else:
        color = _RED
    return PatternFill("solid", fgColor=color)


def export_bucks_list(
    prospects: list[dict],
    output_path: Optional[str] = None,
) -> str:
    """
    Build Buck's outreach Excel workbook.

    prospects: list of dicts with keys from Prospect model + optional
               'contacts' key containing list of contact dicts.

    Returns path to the generated file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")

    if output_path is None:
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        output_path = f"vancon_prospect_list_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Hot Leads (score ≥ 70) ──────────────────────────────────────
    ws_hot = wb.active
    ws_hot.title = "🔥 Hot Leads (70+)"
    hot_prospects = [p for p in prospects if p.get("fit_score", 0) >= 70]
    _write_prospect_sheet(ws_hot, sorted(hot_prospects, key=lambda x: -x.get("fit_score", 0)))

    # ── Sheet 2: All Prospects ────────────────────────────────────────────────
    ws_all = wb.create_sheet("All Prospects")
    _write_prospect_sheet(ws_all, sorted(prospects, key=lambda x: -x.get("fit_score", 0)))

    # ── Sheet 3: Contacts (decision-makers) ──────────────────────────────────
    ws_contacts = wb.create_sheet("Decision Makers")
    _write_contacts_sheet(ws_contacts, prospects)

    # ── Sheet 4: Pipeline stats ───────────────────────────────────────────────
    ws_stats = wb.create_sheet("Stats")
    _write_stats_sheet(ws_stats, prospects)

    path = Path(output_path)
    wb.save(str(path))
    log.info(f"Exported {len(prospects)} prospects to {path}")
    return str(path)


def _write_prospect_sheet(ws, prospects: list[dict]) -> None:
    headers = [
        "Score", "Tier", "Company", "State", "City",
        "Website", "Vista?", "Work Types",
        "Fleet Est.", "Est. ARR",
        "Key Contacts", "Outreach Angle",
        "Pain Point Signals", "Notable Projects",
        "Status", "Scraped At",
    ]

    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDER

    ws.row_dimensions[1].height = 30

    for row_idx, p in enumerate(prospects, 2):
        score = p.get("fit_score", 0)
        tier = p.get("priority_tier", "")
        contacts = p.get("contacts", [])
        contact_str = "; ".join(
            f"{c.get('full_name', '')} ({c.get('title', '')})"
            for c in contacts[:2]
        )

        work_types = p.get("work_types") or []
        if isinstance(work_types, str):
            try:
                work_types = json.loads(work_types)
            except Exception:
                work_types = [work_types]
        work_str = ", ".join(work_types)

        pain_points = p.get("pain_point_signals") or []
        if isinstance(pain_points, str):
            try:
                pain_points = json.loads(pain_points)
            except Exception:
                pain_points = [pain_points]
        pain_str = " | ".join(pain_points[:3]) if pain_points else ""

        projects = p.get("notable_projects") or []
        if isinstance(projects, str):
            try:
                projects = json.loads(projects)
            except Exception:
                projects = [projects]
        projects_str = " | ".join(projects[:2]) if projects else ""

        arr = p.get("estimated_arr", "")
        if isinstance(arr, (int, float)) and arr:
            arr = f"${arr:,.0f}/yr"

        scraped = p.get("scraped_at", "")
        if scraped and "T" in str(scraped):
            scraped = str(scraped)[:10]

        row = [
            score,
            _tier_label(tier),
            p.get("company_name", ""),
            p.get("state", ""),
            p.get("city", ""),
            p.get("website", ""),
            "YES" if p.get("vista_confirmed") else ("maybe" if p.get("uses_vista") else "unknown"),
            work_str,
            p.get("estimated_fleet_size") or p.get("equipment_fleet_est") or "?",
            arr,
            contact_str,
            p.get("outreach_angle", ""),
            pain_str,
            projects_str,
            p.get("status", "identified"),
            scraped,
        ]
        ws.append(row)

        fill = _score_fill(score)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BORDER
            if col_idx == 1:
                cell.fill = fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in (12, 13):  # outreach angle, pain points
                cell.alignment = Alignment(wrap_text=True)

        ws.row_dimensions[row_idx].height = 60 if p.get("outreach_angle") else 20

    # Column widths
    widths = [8, 8, 30, 6, 18, 30, 8, 28, 10, 12, 35, 55, 45, 35, 14, 12]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions


def _write_contacts_sheet(ws, prospects: list[dict]) -> None:
    headers = [
        "Company", "State", "Score", "Vista?",
        "Contact Name", "Title", "Role", "Email", "Phone",
        "LinkedIn", "Decision Maker?", "Email Confidence",
        "Outreach Angle (Company)",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
    ws.row_dimensions[1].height = 25

    row_idx = 2
    for p in sorted(prospects, key=lambda x: -x.get("fit_score", 0)):
        contacts = p.get("contacts", [])
        if not contacts:
            continue
        for c in contacts:
            row = [
                p.get("company_name", ""),
                p.get("state", ""),
                p.get("fit_score", 0),
                "YES" if p.get("vista_confirmed") else "?",
                c.get("full_name", ""),
                c.get("title", ""),
                c.get("role_category", ""),
                c.get("email", ""),
                c.get("phone", ""),
                c.get("linkedin_url", ""),
                "YES" if c.get("is_decision_maker") else "",
                c.get("email_confidence", 0),
                p.get("outreach_angle", ""),
            ]
            ws.append(row)
            cell = ws.cell(row=row_idx, column=3)
            cell.fill = _score_fill(p.get("fit_score", 0))
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = _BORDER
            ws.row_dimensions[row_idx].height = 50
            row_idx += 1

    widths = [28, 6, 8, 8, 25, 30, 18, 32, 16, 45, 14, 15, 55]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions


def _write_stats_sheet(ws, prospects: list[dict]) -> None:
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    def row(label, value, bold=False):
        ws.append([label, value])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    total = len(prospects)
    hot = sum(1 for p in prospects if p.get("fit_score", 0) >= 70)
    warm = sum(1 for p in prospects if 40 <= p.get("fit_score", 0) < 70)
    cold = total - hot - warm
    vista_confirmed = sum(1 for p in prospects if p.get("vista_confirmed"))
    enriched = sum(1 for p in prospects if p.get("contacts"))
    total_contacts = sum(len(p.get("contacts", [])) for p in prospects)
    dm_contacts = sum(
        sum(1 for c in p.get("contacts", []) if c.get("is_decision_maker"))
        for p in prospects
    )
    avg_score = sum(p.get("fit_score", 0) for p in prospects) / total if total else 0

    row("VANCON Technologies — FieldBridge Prospect Database", "", bold=True)
    row(f"Generated: {datetime.now().strftime('%B %d, %Y')}", "")
    row("", "")
    row("Total Prospects", total, bold=True)
    row("Hot Leads (score ≥ 70)", hot)
    row("Warm Leads (score 40–69)", warm)
    row("Cold / Disqualified (score < 40)", cold)
    row("", "")
    row("Average Fit Score", f"{avg_score:.1f} / 100")
    row("Vista Confirmed", vista_confirmed)
    row("Companies Enriched (contacts found)", enriched)
    row("Total Decision-Maker Contacts", total_contacts)
    row("Confirmed Decision Makers", dm_contacts)
    row("", "")
    row("PIPELINE VALUE ESTIMATE", "", bold=True)

    arr_tiers = {
        "Tier 1 ($120K ARR)": sum(1 for p in prospects if p.get("priority_tier") in ("tier_1", "TIER_1")),
        "Tier 2 ($60K ARR)": sum(1 for p in prospects if p.get("priority_tier") in ("tier_2", "TIER_2")),
        "Tier 3 ($30K ARR)": sum(1 for p in prospects if p.get("priority_tier") in ("tier_3", "TIER_3")),
    }
    for label, count in arr_tiers.items():
        row(f"  {label}", count)

    t1 = arr_tiers["Tier 1 ($120K ARR)"] * 120_000
    t2 = arr_tiers["Tier 2 ($60K ARR)"] * 60_000
    t3 = arr_tiers["Tier 3 ($30K ARR)"] * 30_000
    tam = t1 + t2 + t3
    pipeline_at_20pct = tam * 0.20
    row("", "")
    row("Total Addressable ARR (if all close)", f"${tam:,.0f}")
    row("20% Close Rate Estimate", f"${pipeline_at_20pct:,.0f}/yr")


def _tier_label(tier: str) -> str:
    mapping = {"tier_1": "T1 🔥", "tier_2": "T2", "tier_3": "T3"}
    return mapping.get((tier or "").lower(), tier or "")
